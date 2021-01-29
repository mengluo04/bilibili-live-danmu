import json
import struct
import time
import threading
import zlib
from collections import namedtuple
from pathlib import Path

import requests
import websocket
from openpyxl import Workbook, load_workbook

HEADER_STRUCT = struct.Struct('>I2H2I')
# 大端 >
# 四字节长度 I
# 2字节长度 H

HeaderTuple = namedtuple('HeaderTuple', ('pack_len', 'raw_header_size', 'ver', 'operation', 'seq_id'))


class Spider(object):
    def __init__(self):
        self.ws = websocket.create_connection('wss://broadcastlv.chat.bilibili.com:443/sub')
        self.uid = 10
        self.room_id = 764155

    #     url上的地址可能为短id，需要解析为真实的房间id
    def get_real_room_id(self):
        res = requests.get(
            'https://api.live.bilibili.com/xlive/web-room/v1/index/getInfoByRoom?room_id={}'.format(self.room_id))
        data = res.json()
        try:
            room_info = data['data']['room_info']
            room_id = room_info['room_id']
            return room_id
        except Exception as e:
            print(e)
            return self.room_id

    # 解包
    def _handle_message(self, data):
        offset = 0
        while offset < len(data):
            try:
                header = HeaderTuple(*HEADER_STRUCT.unpack_from(data, offset))
            except struct.error:
                break
            if header.operation == 5:
                # 弹幕消息
                body = data[offset + HEADER_STRUCT.size: offset + header.pack_len]
                if header.ver == 2:
                    # zlib加密消息
                    body = zlib.decompress(body)
                    self._handle_message(body)
                else:
                    # 正常json消息
                    try:
                        body = json.loads(body.decode('utf-8'))
                        # 弹幕
                        if body['cmd'] == 'DANMU_MSG':
                            info = body['info']
                            # info结构是数组，第二个值为弹幕内容，第三个为用户信息，包括uid和昵称，第四个为粉丝牌信息，其他不清楚
                            user = info[2][1]
                            msg = info[1]
                            print(user, '：', msg)
                            self.writeExcel(user, msg)
                        #     礼物
                        if body['cmd'] == 'SEND_GIFT':
                            data = body['data']
                            user = data['uname']
                            num = data['num']
                            giftName = data['giftName']
                            action = data['action']
                            print(user, '\t', action, '\t', giftName, '\t', num, '个')
                    except Exception as e:
                        print(e)
                        raise
            else:
                pass
            offset += header.pack_len

    # 封包
    def make_packet(self, data, operation):
        body = json.dumps(data).encode('utf-8')
        header = HEADER_STRUCT.pack(
            HEADER_STRUCT.size + len(body),
            HEADER_STRUCT.size,
            1,
            operation,
            1
        )
        return header + body

    # 登录
    def login(self):
        auth_params = {
            'uid': self.uid,
            'roomid': self.get_real_room_id(),
            'protover': 2,
            'type': 2,
            'platform': 'web',
            'clientver': '2.6.2',
        }
        try:
            self.ws.send(self.make_packet(auth_params, 7))
            print('登陆成功')
        except Exception as e:
            print('登录失败')
            exit(1)

    # 获取消息
    def get_msg(self):
        print('开始接收消息')
        while True:
            try:
                msg_bytes = self.ws.recv()
                self._handle_message(msg_bytes)
            except Exception as e:
                print(e)
                exit(1)

    # 心跳连接
    def keep_alive(self):
        """
        客户端每隔 30 秒发送心跳信息给弹幕服务器
        """
        while True:
            try:
                print('心跳')
                self.ws.send(self.make_packet({}, 2))
                time.sleep(30)
            except Exception as e:
                print(e)
                exit(1)

    # 保存到excel
    def writeExcel(self, nick, msg):
        excel = Path(str(self.get_real_room_id()) + '.xlsx').exists()
        # 文件存在，直接打开
        if excel:
            wb = load_workbook(str(self.get_real_room_id()) + '.xlsx')
            sheet = wb.active
        # 不存在，新建
        else:
            wb = Workbook()
            sheet = wb.active
            sheet.append(['昵称', '弹幕', '发送时间'])
        sheet.append([nick, msg, time.strftime('%y-%m-%d %H:%M:%S')])
        wb.save(str(self.get_real_room_id()) + '.xlsx')


if __name__ == '__main__':
    dm = Spider()
    dm.login()
    t1 = threading.Thread(target=dm.get_msg)
    t2 = threading.Thread(target=dm.keep_alive)
    t2.setDaemon(True)
    t1.start()
    t2.start()
